#!/usr/bin/env python3
"""
emit_xlsx.py — WS6.1 / SPEC v2 §6: the packet workbook emitter.

One seeded sort-night model, rendered as the psi-reconciliation-env
packet's hostile workbook family plus ground_truth.json:

    SPR-T_<sortdate>.xlsx            Summary · Work Area Types · Staffing
    SCN_HubSummary_<exportts>.xlsx   title row 1, headers row 2, EMPTY
                                     padding first column (schema-faithful)
    SCN_EmployeeSummary_<exportts>.xlsx  same hostile layout; User ids
                                     rendered float-style ('2456123.0')
    SVQ_<date>.xlsx                  per-belt service scorecard
    ground_truth.json                the grader's contract — never enters
                                     the agent's environment

Trap toggles (SPEC §3; the v2 ship set is T1/T2/T5/T6/T7):
    T1  the SCN export timestamp in the FILENAME is the next-morning export
        run, not the sort date; the data date is buried in the title cell.
    T2  machine rows (sequential-numeric equipment block from the id
        taxonomy) share the employee table with human loaders.
    T5  one shared-load lane's volume is attributed to a single Zone-3
        belt by SVQ while the scan system splits it across three belts in
        three different zones — the zone-3 concentration ratio computed on
        the SVQ denominator diverges from the correct one.
    T6  ID substitution + shared scanner. Every human row carries the
        device convention Scanner ID == EMP<employee id>; machines carry
        none. Planted: (a) one loader's work rides another loader's login
        while keeping the absent loader's device — the invisible loader is
        inferable only by joining the device link with the staffing
        sheet's Outbound Load roster; (b) one pair shares one device with
        honest logins — a mismatch that is NOT misattribution.
    T7  two unresolvable attribution anomalies: a well-formed unrostered
        user id (self-consistent device, no staffing line), and a rostered
        user whose volume splits onto a device registered to nobody.
        Resolution is not derivable from the exports; ground truth lists
        them so the grader can check the memo flags without fabricating.

Everything derives from the NGATE world: calibration.json lognormals and
dirichlets, the hub profile's zones, the id taxonomy's fictional blocks.
Anomaly counts come from calibration anomaly_rates (documented failure
modes, never corpus-fit). Deterministic for (seed, sortdate, traps): fixed
workbook properties, fixed zip member timestamps.

USAGE
    python3 emit_xlsx.py --out-dir instance/ --seed 7 --sortdate 2026-06-24
    python3 emit_xlsx.py ... --traps T1,T2,T5      (default: the ship set)
"""
import argparse
import datetime as dt
import io
import json
import random
import re
import sys
import zipfile
from pathlib import Path

import openpyxl

GEN = Path(__file__).resolve().parent
CALIBRATION = GEN / "calibration/calibration.json"
NGATE_PROFILE = GEN / "world/hub_profile.northgate.json"

FIXED_STAMP = dt.datetime(2026, 1, 1, 0, 0, 0)
ALL_TRAPS = ("T1", "T2", "T5", "T6", "T7")

WORK_AREA_TYPES = ("Unload", "Primary", "Outbound Load", "Smalls Sort",
                   "Air Recovery", "Pickoff")


def _belt_key(profile_belt):
    """profile 'PD-01' → calibration dirichlet key 'PD-1'."""
    return "PD-" + str(int(profile_belt.split("-")[1]))


def build_night(calib, profile, seed, sortdate, traps):
    rng = random.Random(f"{seed}:{sortdate}")
    date = dt.date.fromisoformat(sortdate)
    dow = date.weekday()
    if dow > 4:
        raise SystemExit(f"emit_xlsx: {sortdate} is a weekend — Twilight runs Mon–Fri")

    belts = list(profile["topology"]["outbound_pd_belts"])
    zones = {z["name"]: list(z["belts"]) for z in profile["topology"]["outbound_zones"]}
    z3_belts = zones["Zone 3"]

    # volume level + belt shares from the world's own fits
    vp = calib["volume_params"][str(dow)]
    volume = round(rng.lognormvariate(vp["mu_ln"], vp["sigma_ln"]))
    alpha = calib.get("dirichlet_alpha_by_dow", {}).get(str(dow)) or calib["dirichlet_alpha"]
    keys = [_belt_key(b) for b in belts] + ["AIRSORT", "SMALLS"]
    draws = {k: rng.gammavariate(max(alpha.get(k, 1.0), 1e-3), 1.0) for k in keys}
    total_draw = sum(draws.values())

    # reserve the shared-load lane volume, then distribute the rest
    shared_v = round(volume * rng.uniform(0.05, 0.07))
    base_pool = volume - shared_v
    belt_vol = {}
    for b in belts:
        belt_vol[b] = round(base_pool * draws[_belt_key(b)] / total_draw)
    air_vol = round(base_pool * draws["AIRSORT"] / total_draw)
    smalls_vol = round(base_pool * draws["SMALLS"] / total_draw)
    # rounding residue → largest belt
    residue = base_pool - sum(belt_vol.values()) - air_vol - smalls_vol
    belt_vol[max(belt_vol, key=belt_vol.get)] += residue

    # Basis translation (locked decision, WS3.5 pattern): rescale the zone-3
    # belts so the night's concentration ratio realizes the profile's own
    # κ(dow) — the packet instance, the tracker demo, and the contract then
    # tell one consistent story. Non-z3 belts absorb the complement.
    dow_name = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")[dow]
    kappa = profile["metric_contract"]["structural_constants"]["kappa_dow_steady"][dow_name]
    anchor_third = shared_v - 2 * (shared_v // 3)
    z3_base_target = round(kappa * volume) - anchor_third
    z3_base = sum(belt_vol[b] for b in z3_belts)
    other_base = sum(v for b, v in belt_vol.items() if b not in z3_belts)
    f3 = z3_base_target / z3_base
    f_other = (other_base + z3_base - z3_base_target) / other_base
    for b in belt_vol:
        belt_vol[b] = round(belt_vol[b] * (f3 if b in z3_belts else f_other))
    residue = (base_pool - air_vol - smalls_vol) - sum(belt_vol.values())
    belt_vol[max((b for b in belt_vol if b not in z3_belts),
                 key=lambda b: belt_vol[b])] += residue

    # T5 topology: scan splits the shared lane across one belt per zone;
    # SVQ bundles the whole lane onto the Zone-3 anchor.
    anchor = rng.choice(z3_belts)
    split = [anchor, rng.choice(zones["Zone 2"]), rng.choice(zones["Zone 1"])]
    scan_belt_vol = dict(belt_vol)
    svq_belt_vol = dict(belt_vol)
    thirds = [shared_v // 3, shared_v // 3, shared_v - 2 * (shared_v // 3)]
    if "T5" in traps:
        for b, v in zip(split, thirds):
            scan_belt_vol[b] += v
        svq_belt_vol[anchor] += shared_v
    else:
        for b, v in zip(split, thirds):
            scan_belt_vol[b] += v
            svq_belt_vol[b] += v

    # Denominator semantics (the heart of T5): the CORRECT ratio is zone-3
    # scan volume over the scan system's full hub total (outbound belts +
    # air + smalls — production κ semantics); the NAIVE ratio is what SVQ
    # alone yields — bundled attribution over its outbound-only coverage.
    def z3_ratio(vols, denom):
        return round(sum(vols[b] for b in z3_belts) / denom, 4)

    # roster (clean humans) + machine block (T2)
    roster_size = max(calib["roster"]["size_min"],
                      min(calib["roster"]["size_max"],
                          round(rng.gauss(calib["roster"]["size_mean"],
                                          calib["roster"]["size_sd"]))))
    lo, hi = calib["id_taxonomy"]["human_id"]["block"]
    roster_ids = rng.sample(range(lo, hi + 1), roster_size)
    roster = [{"name": f"Employee-{i + 1}", "id": roster_ids[i],
               "ft": rng.random() < calib["roster"]["ft_share"]}
              for i in range(roster_size)]

    machine_ids = []
    if "T2" in traps:
        m_lo, m_hi = calib["id_taxonomy"]["machine_blocks"]["blocks"][0]
        start = rng.randint(m_lo, m_hi - 8)
        machine_ids = list(range(start, start + rng.randint(3, 6)))

    # loader scan attribution: coverage of each belt's volume, split among
    # 2–5 roster loaders (+ machine rows when T2). Every human row carries
    # its owner's device (EMP<employee id>) until planting rewrites it.
    loader_pool = rng.sample(roster, min(len(roster), 5 * len(belts)))
    scan_rows, pool_i, belt_group = [], 0, {}
    for b in belts:
        n = rng.randint(2, 5)
        loaders = loader_pool[pool_i:pool_i + n]
        pool_i += n
        belt_group[b] = [p["id"] for p in loaders]
        coverage = round(scan_belt_vol[b] * rng.uniform(0.82, 0.92))
        weights = [rng.uniform(0.6, 1.4) for _ in loaders]
        wsum = sum(weights)
        for person, w in zip(loaders, weights):
            scan_rows.append({"user": person["id"], "belt": b,
                              "outbound": round(coverage * w / wsum),
                              "scanner": f"EMP{person['id']}"})
    for i, mid in enumerate(machine_ids):
        b = belts[i % len(belts)]
        scan_rows.append({"user": mid, "belt": b,
                          "outbound": round(volume * rng.uniform(0.004, 0.012)),
                          "scanner": None})

    # staffing coherence: whoever scans outbound is rostered Outbound Load;
    # the outbound headcount realizes calibration outbound_share, padded
    # with non-scanning outbound staff so absence corroborates, never
    # convicts, on its own.
    used = {p["id"] for p in loader_pool[:pool_i]}
    others = [p["id"] for p in roster if p["id"] not in used]
    pad = rng.sample(others, max(0, round(calib["roster"]["outbound_share"]
                                          * len(roster)) - len(used)))
    outbound_staff = used | set(pad)
    other_areas = [a for a in WORK_AREA_TYPES if a != "Outbound Load"]
    staffing_area, oa_i = {}, 0
    for p in roster:
        if p["id"] in outbound_staff:
            staffing_area[p["id"]] = "Outbound Load"
        else:
            staffing_area[p["id"]] = other_areas[oa_i % len(other_areas)]
            oa_i += 1

    # T6/T7 planting: parties and belts drawn unconditionally (one story
    # per belt, all parties distinct by pool-slice disjointness), applied
    # per toggle. Planting rewrites attribution — it never invents volume,
    # so per-belt accounting matches the hub summary regardless of traps.
    b_sub, b_vessel, b_shared, b_orphan, b_udev = rng.sample(belts, 5)
    s_id = rng.choice(belt_group[b_sub])            # the invisible loader
    v_id = rng.choice(belt_group[b_vessel])         # the credited vessel
    p_id, q_id = rng.sample(belt_group[b_shared], 2)  # device owner + rider
    donor_id = rng.choice(belt_group[b_orphan])
    t_id = rng.choice(belt_group[b_udev])

    def _fresh_block_id(taken):
        lo, hi = calib["id_taxonomy"]["human_id"]["block"]
        while True:
            cand = rng.randint(lo, hi)
            if cand not in taken:
                return cand
    roster_id_set = {p["id"] for p in roster}
    orphan_id = _fresh_block_id(roster_id_set)
    udev_digits = _fresh_block_id(roster_id_set | {orphan_id})

    def _row(user):
        return next(r for r in scan_rows if r["user"] == user)

    def _split(vol, frac):
        head = round(vol * frac)
        return head, vol - head

    t6_info = {"scanner_convention": "EMP<7-digit employee id>",
               "substitution": None, "shared_scanner": None,
               "anomalous_records": []}
    if "T6" in traps:
        # (a) ID substitution: S's work rides V's login, on S's device.
        s_row = _row(s_id)
        scan_rows.remove(s_row)
        a, b_ = _split(s_row["outbound"], rng.uniform(0.55, 0.7))
        sub_records = [{"user": v_id, "belt": b_sub,
                        "scanner": f"EMP{s_id}", "outbound": x}
                       for x in (a, b_)]
        scan_rows.extend(dict(r) for r in sub_records)
        # (b) shared scanner: Q logs honestly but rides P's device.
        q_row = _row(q_id)
        q_row["scanner"] = f"EMP{p_id}"
        shared_records = [{"user": q_id, "belt": b_shared,
                           "scanner": f"EMP{p_id}",
                           "outbound": q_row["outbound"]}]
        t6_info.update(
            substitution={"invisible_loader_id": s_id,
                          "credited_user_id": v_id, "belt": b_sub,
                          "device": f"EMP{s_id}", "records": sub_records},
            shared_scanner={"device_owner_id": p_id, "second_user_id": q_id,
                            "belt": b_shared, "device": f"EMP{p_id}",
                            "records": shared_records},
            anomalous_records=sub_records + shared_records)

    t7_info = {"unresolvables": []}
    if "T7" in traps:
        # (a) unrostered user: real volume carved from the donor's rows,
        # self-consistent device, no staffing line to resolve against.
        d_row = _row(donor_id)
        carve, keep = _split(d_row["outbound"], rng.uniform(0.25, 0.4))
        d_row["outbound"] = keep
        a, b_ = _split(carve, rng.uniform(0.4, 0.6))
        orphan_records = [{"user": orphan_id, "belt": b_orphan,
                           "scanner": f"EMP{orphan_id}", "outbound": x}
                          for x in (a, b_)]
        scan_rows.extend(dict(r) for r in orphan_records)
        # (b) unregistered device: T's volume splits onto a device that
        # maps to no roster id — provenance not derivable.
        t_row = _row(t_id)
        own, foreign = _split(t_row["outbound"], rng.uniform(0.45, 0.6))
        t_row["outbound"] = own
        udev_records = [{"user": t_id, "belt": b_udev,
                         "scanner": f"EMP{udev_digits}", "outbound": foreign}]
        scan_rows.extend(dict(r) for r in udev_records)
        t7_info["unresolvables"] = [
            {"kind": "unrostered_user", "user_id": orphan_id,
             "belt": b_orphan, "records": orphan_records,
             "why": "id is well-formed and in the human block but appears "
                    "nowhere in staffing; the exports carry no identity "
                    "source to resolve it — the fix is data collection"},
            {"kind": "unregistered_device", "user_id": t_id,
             "device": f"EMP{udev_digits}", "belt": b_udev,
             "records": udev_records,
             "why": "the second device's digits match no rostered id; "
                    "whether the volume is really this loader's cannot be "
                    "established from the exports"}]

    rng.shuffle(scan_rows)

    # cross-source totals: SPR actual differs from the scan basis by a
    # small real gap; SVQ moves volume between belts, never the total
    spr_actual = round(volume * (1 + rng.uniform(-0.004, 0.004)))
    spr_planned = round(volume * (1 + rng.uniform(-0.03, 0.05)))
    scn_total = sum(scan_belt_vol.values()) + air_vol + smalls_vol
    svq_total = sum(svq_belt_vol.values())
    z3_correct = z3_ratio(scan_belt_vol, scn_total)
    z3_naive = z3_ratio(svq_belt_vol, svq_total)

    export_dt = (dt.datetime.combine(date, dt.time(22, 10)) if "T1" not in traps
                 else dt.datetime.combine(date + dt.timedelta(days=1),
                                          dt.time(6, rng.randint(2, 49))))

    gap = lambda a, b: round(100.0 * (a - b) / b, 3)
    return {
        "sortdate": sortdate, "dow": dow, "volume": volume,
        "belts": belts, "zones": zones,
        "scan_belt_vol": scan_belt_vol, "svq_belt_vol": svq_belt_vol,
        "air_vol": air_vol, "smalls_vol": smalls_vol,
        "roster": roster, "scan_rows": scan_rows,
        "machine_ids": machine_ids,
        "staffing_area": staffing_area,
        "t6": t6_info, "t7": t7_info,
        "shared": {"volume": shared_v, "anchor": anchor, "split": split,
                   "thirds": thirds},
        "z3_correct": z3_correct, "z3_naive": z3_naive,
        "spr_actual": spr_actual, "spr_planned": spr_planned,
        "scn_total": scn_total, "svq_total": svq_total,
        "gaps": {"spr_vs_scn": gap(spr_actual, scn_total),
                 "svq_vs_scn": gap(svq_total, scn_total)},
        "export_dt": export_dt,
        "rng": rng,
    }


# ── workbook writers ──────────────────────────────────────────────────

def _fresh_wb():
    wb = openpyxl.Workbook()
    wb.properties.created = FIXED_STAMP
    wb.properties.modified = FIXED_STAMP
    return wb


def _save_deterministic(wb, path):
    """Save with fixed zip member timestamps so identical content is
    identical bytes. docProps/core.xml needs its dates re-pinned here:
    openpyxl overwrites dcterms:modified with utcnow() AT SAVE, so the
    _fresh_wb stamp survives only for same-second rebuilds."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    out = io.BytesIO()
    with zipfile.ZipFile(buf) as zin, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            info = zipfile.ZipInfo(item.filename,
                                   date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            data = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                data = re.sub(rb'(type="dcterms:W3CDTF">)[^<]+',
                              rb"\g<1>2026-01-01T00:00:00Z", data)
            zout.writestr(info, data)
    path.write_bytes(out.getvalue())


def write_spr(night, out_dir):
    wb = _fresh_wb()
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Planned", "Values", "Actual"),
        ("18:00", "Sort Start", "18:00"),
        ("21:30", "Sort Down", "21:40"),
        (night["spr_planned"], "Volume", night["spr_actual"]),
        (round(night["spr_planned"] / 125.0, 1), "Paid Hours",
         round(night["spr_actual"] / 124.2, 1)),
    ]
    for r in rows:
        ws.append(r)

    wa = wb.create_sheet("Work Area Types")
    wa.append(("Work Area Type", "Work Area", "Work Area Job", "Employee Name",
               "Employee ID", "Planned % Vol", "Actual % Vol"))
    rng = night["rng"]
    for area in WORK_AREA_TYPES:
        share = round(rng.uniform(0.03, 0.30), 4)
        wa.append((area, "", "", "", "", share,
                   round(share * rng.uniform(0.9, 1.1), 4)))

    st = wb.create_sheet("Staffing")
    st.append(("Work Area Type", "Work Area", "Work Area Job", "Employee Name",
               "Employee ID", "FT/PT", "Pay Guarantee", "Pre-Start"))
    for person in night["roster"]:
        area = night["staffing_area"][person["id"]]
        st.append((area, f"{area[:3].upper()}-{person['id'] % 7 + 1}", area,
                   person["name"], str(person["id"]),
                   "FT" if person["ft"] else "PT", 0, 0))
    path = out_dir / f"SPR-T_{night['sortdate']}.xlsx"
    _save_deterministic(wb, path)
    return path.name


def write_scn_hub(night, out_dir):
    ts = night["export_dt"].strftime("%Y%m%d_%H%M")
    wb = _fresh_wb()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append((None, f"Hub Summary — Data Date: {night['sortdate']}"))
    ws.append((None, "Belt", "LooseOutbound", "LooseInbound", "ULDsInbound",
               "In Building", "BagsCreated", "BagsLinked", "Packages inLinkedBags"))
    rng = night["rng"]

    def row(belt, vol):
        bags = round(vol * rng.uniform(0.05, 0.09))
        loose = vol - bags
        ws.append((None, belt, loose, 0, 0, rng.randint(0, 40),
                   bags + rng.randint(0, 8), bags,
                   round(bags * rng.uniform(11, 15))))

    for b in night["belts"]:
        row(b, night["scan_belt_vol"][b])
    row("AIR-1", night["air_vol"])
    row("SMALLS", night["smalls_vol"])
    path = out_dir / f"SCN_HubSummary_{ts}.xlsx"
    _save_deterministic(wb, path)
    return path.name


def write_scn_employee(night, out_dir):
    ts = night["export_dt"].strftime("%Y%m%d_%H%M")
    wb = _fresh_wb()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append((None, "Employee Summary", None, None,
               "Volume Attribution", None, None, "Device Link"))
    ws.append((None, "User", "Belt", "Total Packages", "Inbound Volume",
               "Outbound Volume", "Keyed Packages", "Scanner ID"))
    for r in night["scan_rows"]:
        keyed = round(r["outbound"] * night["rng"].uniform(0.0, 0.02))
        ws.append((None, f"{r['user']}.0", r["belt"], float(r["outbound"] + keyed),
                   0.0, float(r["outbound"]), float(keyed), r["scanner"]))
    path = out_dir / f"SCN_EmployeeSummary_{ts}.xlsx"
    _save_deterministic(wb, path)
    return path.name


def write_svq(night, out_dir):
    wb = _fresh_wb()
    ws = wb.active
    ws.title = "Scorecard"
    ws.append(("Zone", "Belt", "Volume", "Exceptions", "Service %"))
    rng = night["rng"]
    zone_of = {b: z for z, bs in night["zones"].items() for b in bs}
    for b in night["belts"]:
        vol = night["svq_belt_vol"][b]
        exc = round(vol * rng.uniform(0.001, 0.006))
        ws.append((zone_of[b], b, vol, exc,
                   round(100 * (1 - exc / max(vol, 1)), 2)))
    path = out_dir / f"SVQ_{night['sortdate']}.xlsx"
    _save_deterministic(wb, path)
    return path.name


def ground_truth(night, traps, seed, files):
    return {
        "seed": seed, "sortdate": night["sortdate"],
        "traps_enabled": sorted(traps),
        "data_date": night["sortdate"],
        "files": files,
        "roster_size": len(night["roster"]),
        "roster_ids": sorted(p["id"] for p in night["roster"]),
        "totals": {"spr_actual": night["spr_actual"], "scn": night["scn_total"],
                   "svq": night["svq_total"]},
        "gaps_pct": night["gaps"],
        "t1": {"filename_timestamp": night["export_dt"].strftime("%Y%m%d_%H%M"),
               "export_ran": night["export_dt"].isoformat(sep=" ")},
        "t2": {"machine_ids": night["machine_ids"]},
        "t5": {"shared_lane_volume": night["shared"]["volume"],
               "svq_anchor_belt": night["shared"]["anchor"],
               "scan_split_belts": night["shared"]["split"],
               "scan_split_volumes": night["shared"]["thirds"]},
        "t6": night["t6"],
        "t7": night["t7"],
        "zone3": {"belts": night["zones"]["Zone 3"],
                  "correct": night["z3_correct"],
                  "naive_from_svq": night["z3_naive"],
                  "note": "correct = zone-3 scan volume over the scan hub total "
                          "(outbound + air + smalls, production kappa semantics); "
                          "naive = SVQ's bundled attribution over its "
                          "outbound-only coverage"},
    }


def main(argv=None):
    ap = argparse.ArgumentParser(description="Emit one packet workbook instance.")
    ap.add_argument("--out-dir", type=Path, required=True)
    ap.add_argument("--seed", type=int, default=7)
    ap.add_argument("--sortdate", default="2026-06-24")
    ap.add_argument("--traps", default=",".join(ALL_TRAPS),
                    help=f"comma list from {','.join(ALL_TRAPS)} "
                         "(default: the full ship set)")
    args = ap.parse_args(argv)

    traps = {t.strip().upper() for t in args.traps.split(",") if t.strip()}
    unknown = traps - set(ALL_TRAPS)
    if unknown:
        raise SystemExit(f"emit_xlsx: unknown traps {sorted(unknown)} "
                         f"(this build ships {ALL_TRAPS})")

    calib = json.loads(CALIBRATION.read_text())
    profile = json.loads(NGATE_PROFILE.read_text())
    night = build_night(calib, profile, args.seed, args.sortdate, traps)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    files = [write_spr(night, args.out_dir),
             write_scn_hub(night, args.out_dir),
             write_scn_employee(night, args.out_dir),
             write_svq(night, args.out_dir)]
    gt = ground_truth(night, traps, args.seed, files)
    (args.out_dir / "ground_truth.json").write_text(json.dumps(gt, indent=2) + "\n")
    print(f"emit_xlsx: {night['sortdate']} instance (seed {args.seed}, "
          f"traps {'+'.join(sorted(traps)) or 'none'}, volume {night['volume']:,}) "
          f"→ {args.out_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
